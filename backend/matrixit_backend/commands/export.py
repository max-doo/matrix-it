"""
Export 相关命令模块。

包含：export_excel, export_pdfs
所有重型依赖（openpyxl）均采用懒加载。
"""
import json
import os
from pathlib import Path
from typing import Dict, List


def export_excel(
    db_path: str,
    root_dir: str,
    config_path: str,
    fields_path: str,
    keys: List[str],
    output_path: str,
    filename: str,
) -> dict:
    """
    [IPC 命令] 导出 Excel 文件。
    
    仅导出 processed_status == 'done' 的条目，按飞书同步字段顺序构建列。
    
    Args:
        keys: 需要导出的 item_key 列表
        output_path: 导出目录路径
        filename: 文件名 (含 .xlsx 后缀)
        
    Returns:
        Dict: {written, skipped, output_path, failures[]}
    """
    import sys
    
    # 懒加载 openpyxl
    try:
        from openpyxl import Workbook
    except ImportError as e:
        return {
            "written": 0, 
            "skipped": 0, 
            "output_path": "", 
            "failures": [f"openpyxl 导入失败: {e}", f"sys.path: {sys.path}"]
        }

    # 懒加载其他依赖
    from matrixit_backend import feishu, storage
    from matrixit_backend.config import load_config

    config = load_config(config_path)
    items_idx = storage.get_items_index(db_path)
    
    # 加载字段定义
    fields_def: dict = {}
    if isinstance(config.get("fields"), dict):
        fields_def = config.get("fields", {})
    else:
        try:
            with open(fields_path, "r", encoding="utf-8") as f:
                fields_def = json.load(f)
        except Exception:
            fields_def = {}
    
    fc = config.get("feishu", {}) if isinstance(config.get("feishu"), dict) else {}
    
    def _unique_keep_order(seq: List[str]) -> List[str]:
        seen = set()
        out: List[str] = []
        for x in seq:
            s = str(x or "").strip()
            if not s or s in seen:
                continue
            seen.add(s)
            out.append(s)
        return out
    
    meta_sync_raw = fc.get("meta_sync")
    if not isinstance(meta_sync_raw, list):
        meta_sync_raw = fc.get("metaSync") if isinstance(fc.get("metaSync"), list) else None
    enabled_meta: set = set()
    if isinstance(meta_sync_raw, list):
        enabled_meta = set([str(x).strip() for x in meta_sync_raw if str(x).strip()])
    else:
        enabled_meta = set([k for k in feishu.FEISHU_META_DEFAULT_ORDER if k != "doi"])
    
    def _should_sync_meta_key(k: str) -> bool:
        if k in feishu.FEISHU_META_FIXED_KEYS:
            return True
        return k in enabled_meta
    
    ui_cfg = config.get("ui", {}) if isinstance(config.get("ui", {}), dict) else {}
    tc = ui_cfg.get("table_columns", {}) if isinstance(ui_cfg.get("table_columns", {}), dict) else {}
    matrix = tc.get("matrix", {}) if isinstance(tc.get("matrix", {}), dict) else {}
    analysis_ui = matrix.get("analysis", {}) if isinstance(matrix.get("analysis", {}), dict) else {}
    analysis_order_raw = analysis_ui.get("order")
    analysis_order = _unique_keep_order(analysis_order_raw if isinstance(analysis_order_raw, list) else [])
    
    mapping: Dict[str, str] = {}
    used_names: Dict[str, str] = {}
    
    def _add_mapping(section: str, key: str, v: dict) -> None:
        name = str(v.get("name") or "").strip()
        if not name:
            return
        if name in used_names and used_names[name] != str(key):
            return
        used_names[name] = str(key)
        mapping[str(key)] = name
    
    meta_defs = fields_def.get("meta_fields", {}) if isinstance(fields_def.get("meta_fields", {}), dict) else {}
    analysis_defs = fields_def.get("analysis_fields", {}) if isinstance(fields_def.get("analysis_fields", {}), dict) else {}
    
    meta_keys_ordered = _unique_keep_order(feishu.FEISHU_META_DEFAULT_ORDER + list(meta_defs.keys()))
    for k in meta_keys_ordered:
        if k not in meta_defs:
            continue
        if not _should_sync_meta_key(k):
            continue
        v = meta_defs.get(k)
        if isinstance(v, dict):
            _add_mapping("meta_fields", k, v)
    
    analysis_keys_ordered = _unique_keep_order(analysis_order + list(analysis_defs.keys()))
    for k in analysis_keys_ordered:
        if k not in analysis_defs:
            continue
        v = analysis_defs.get(k)
        if isinstance(v, dict):
            _add_mapping("analysis_fields", k, v)
    
    ordered_json_keys: List[str] = []
    seen_keys = set()
    for k in meta_keys_ordered:
        if k in mapping and k not in seen_keys:
            ordered_json_keys.append(k)
            seen_keys.add(k)
    for k in analysis_keys_ordered:
        if k in mapping and k not in seen_keys:
            ordered_json_keys.append(k)
            seen_keys.add(k)
    
    req_keys = set([str(k).strip() for k in (keys or []) if str(k).strip()])
    targets: List[dict] = []
    skipped = 0
    for k, it in items_idx.items():
        if req_keys and str(k) not in req_keys:
            continue
        if it.get("processed_status") != "done":
            skipped += 1
            continue
        targets.append(it)
    
    if not targets:
        return {"written": 0, "skipped": skipped, "output_path": "", "failures": []}
    
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "文献导出"
    
    headers = [mapping.get(k, k) for k in ordered_json_keys]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    def _get_by_path(obj: object, path: str) -> object:
        cur: object = obj
        for part in str(path).split("."):
            if not isinstance(cur, dict):
                return None
            cur = cur.get(part)
        return cur
    
    def _format_value(val: object) -> object:
        if val is None:
            return ""
        if isinstance(val, (list, tuple)):
            return ", ".join([str(x).strip() for x in val if str(x).strip()])
        if isinstance(val, dict):
            return json.dumps(val, ensure_ascii=False)
        return val
    
    failures: List[str] = []
    for row_idx, item in enumerate(targets, 2):
        for col_idx, jk in enumerate(ordered_json_keys, 1):
            val: object = None
            if "." in jk:
                val = _get_by_path(item, jk)
            elif jk in item:
                val = item.get(jk)
            elif jk == "tags":
                meta = item.get("meta_extra") if isinstance(item.get("meta_extra"), dict) else {}
                val = meta.get("tags") if isinstance(meta, dict) else None
            elif jk == "collections" and isinstance(item.get("collections"), list):
                val = [c.get("name") for c in item.get("collections", []) if isinstance(c, dict) and c.get("name")]
            
            if jk == "type":
                val = feishu.map_literature_type_to_zh(val)
            
            ws.cell(row=row_idx, column=col_idx, value=_format_value(val))
    
    out_dir = Path(output_path)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / filename
    try:
        wb.save(str(out_file))
    except Exception as e:
        failures.append(f"保存文件失败: {e}")
        return {"written": 0, "skipped": skipped, "output_path": "", "failures": failures}
    
    return {"written": len(targets), "skipped": skipped, "output_path": str(out_file), "failures": failures}


def export_pdfs(
    db_path: str,
    root_dir: str,
    config_path: str,
    keys: List[str],
    output_dir: str,
) -> dict:
    """
    [IPC 命令] 导出 PDF 附件。
    
    按集合树结构（使用 path 字段）复制 PDF 到目标目录。
    
    Args:
        keys: 需要导出的 item_key 列表
        output_dir: 导出目录路径
        
    Returns:
        Dict: {exported, skipped_no_pdf, skipped_missing, output_dir, failures[]}
    """
    import shutil
    import re
    
    # 懒加载重型依赖
    from matrixit_backend import storage, zotero
    from matrixit_backend.config import load_config
    
    config = load_config(config_path)
    items_idx = storage.get_items_index(db_path)
    zotero_dir = zotero.get_zotero_dir(config)
    
    req_keys = set([str(k).strip() for k in (keys or []) if str(k).strip()])
    
    exported = 0
    skipped_no_pdf = 0
    skipped_missing = 0
    failures: List[str] = []
    
    def _sanitize_path_part(name: str) -> str:
        return re.sub(r'[\\/:*?"<>|]', '_', str(name or "").strip()).rstrip(" .")
    
    def _sanitize_full_path(path: str) -> str:
        """清洗完整路径，保留分隔符但清洗每个部分"""
        parts = str(path or "").split("/")
        sanitized_parts = [_sanitize_path_part(p) for p in parts if p.strip()]
        return "/".join(sanitized_parts)
    
    out_base = Path(output_dir)
    out_base.mkdir(parents=True, exist_ok=True)
    
    used_paths: Dict[str, int] = {}
    
    for k, it in items_idx.items():
        if req_keys and str(k) not in req_keys:
            continue
        
        pdf_path = zotero.resolve_pdf_path(it, zotero_dir, base_dir=str(root_dir))
        if not pdf_path:
            skipped_no_pdf += 1
            continue
        
        if not os.path.exists(pdf_path):
            skipped_missing += 1
            failures.append(f"{k}: PDF 文件不存在 ({pdf_path})")
            continue
        
        collections = it.get("collections", [])
        sub_dir = ""
        if isinstance(collections, list) and collections:
            first_col = collections[0] if collections else {}
            if isinstance(first_col, dict):
                col_path = first_col.get("path") or ""
                if col_path:
                    sub_dir = _sanitize_full_path(col_path)
                else:
                    sub_dir = _sanitize_path_part(first_col.get("name") or "")
        
        if sub_dir:
            target_dir = out_base
            for part in sub_dir.split("/"):
                if part:
                    target_dir = target_dir / part
        else:
            target_dir = out_base
        
        target_dir.mkdir(parents=True, exist_ok=True)
        
        src_filename = os.path.basename(pdf_path)
        target_file = target_dir / src_filename
        
        target_str = str(target_file)
        if target_str in used_paths:
            used_paths[target_str] += 1
            stem = target_file.stem
            suffix = target_file.suffix
            target_file = target_dir / f"{stem}_{used_paths[target_str]}{suffix}"
        else:
            used_paths[target_str] = 0
        
        try:
            shutil.copy2(pdf_path, str(target_file))
            exported += 1
        except Exception as e:
            failures.append(f"{k}: 复制失败 ({e})")
    
    return {
        "exported": exported,
        "skipped_no_pdf": skipped_no_pdf,
        "skipped_missing": skipped_missing,
        "output_dir": str(out_base),
        "failures": failures,
    }
